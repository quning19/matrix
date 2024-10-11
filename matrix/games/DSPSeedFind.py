#coding=utf-8

import os
import csv
import pandas as pd
from openpyxl import Workbook
import openpyxl.workbook
from matrix.base.BaseJob import BaseJob
from matrix.utils import get_all_file_list


class DSPSeedFind(BaseJob):

    property_list = [{'name':'seed', 'field':"星系名字"},
                     {'name':'star', 'field':"星区数量"},
                     {'name':'ocount', 'field':"潮汐锁定O星数"},
                     {'name':'ol1', 'field':"O星亮度 1"},
                     {'name':'od1', 'field':"O星距离 1"},
                     {'name':'ol2', 'field':"O星亮度 2"},
                     {'name':'od2', 'field':"O星距离 2"},
                    #  {'name':'star', 'field':"星区数量"},
                     ]

    def __init__(self, options):
        BaseJob.__init__(self, options)
        self._data_list = []
        pass

    def run(self):
        self.logger.info('start')

        all_file_list = []
        get_all_file_list(self.get_options('input_path'), all_file_list, '.csv')
        for i in range(len(all_file_list)):
            self.read_single_seed(all_file_list[i])

        self.save_to_file()

    def read_single_seed(self, filename):
        if 'single' not in filename:
            return
        data = {}
        seed, star = self.get_seed_star(filename)
        data['seed'] = str(seed)
        data['star'] = str(star)

        self.load_detial(filename, data)
        self._data_list.append(data)

    def load_detial(self, filename, data):
        
        df = pd.read_csv(filename, index_col=False)
        # print(df.columns)
        ostar_count = 0
        for i in range(len(df.index)):
            star_type = df.at[i,'星系类型']
            if star_type != 'O型恒星':
                continue
            planet_type = df.at[i,'星球类型']
            
            if '永昼永夜' in planet_type and (planet_type.find('永昼永夜') > planet_type.find(';') or planet_type.count('永昼永夜') > 1):
                ostar_count += 1
                data['od'+str(ostar_count)] = df.at[i,'距离']
                data['ol'+str(ostar_count)] = df.at[i,'亮度']

        data['ocount'] = str(ostar_count)


    def get_seed_star(self, filename):
        short_name = os.path.splitext(os.path.basename(filename))[0]
        arr = short_name.split("_")
        seed = arr[0]
        star = 64 if len(arr) == 2 else arr[1]
        return seed, star
    
    def save_to_file(self):
        file_path = os.path.join(self.get_options('input_path'), 'export.xlsx')
        if os.path.exists(file_path):
            os.remove(file_path)
        
        wb = Workbook()
        ws = wb.active

        for j in range(len(DSPSeedFind.property_list)):
            ws.cell(1, j + 1).value = DSPSeedFind.property_list[j]['field']

        for i in range(len(self._data_list)):
            for j in range(len(DSPSeedFind.property_list)):
                property_name = DSPSeedFind.property_list[j]['name']
                if property_name in self._data_list[i]:
                    ws.cell(i + 2, j + 1).value = str(self._data_list[i][property_name])

        wb.save(file_path)