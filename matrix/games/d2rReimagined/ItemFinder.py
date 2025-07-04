#coding=utf-8

import json
import os
import shutil

import yaml
import pandas as pd
from matrix.games.d2rReimagined import *
from openpyxl.styles import Font, Alignment
from matrix.base.BaseJob import BaseJob


class ItemFinder(BaseJob):

    excel_relative_path = 'data/global/excel/'
    excel_path = None
    work_path = None
    debug_path = None

    config_list = []

    def __init__(self, options):
        BaseJob.__init__(self, options)

        yaml_path = os.path.join(os.path.dirname(__file__), 'D2rrConfig.yaml')

        with open(yaml_path, 'r', encoding='utf-8') as f:
            self.d2r_config = yaml.load(f, yaml.FullLoader)

        self.excel_path = os.path.join(self.d2r_config['mod_d2rr_root'], self.excel_relative_path)
        self.work_path = self.get_options('work_path')
        self.debug_path = os.path.join(self.work_path, 'debug_export')
        if os.path.exists(self.debug_path):
            shutil.rmtree(self.debug_path)
        os.makedirs(self.debug_path)


    def run(self):
        self.logger.info('start')

        if not os.path.exists(self.excel_path):
            self.logger.error(f'Excel directory not found: {self.excel_path}')
            return
        
        item_detail_file_path = os.path.join(self.work_path, 'ItemDetails.xlsx')

        self.logger.info(f'Reading file: {item_detail_file_path}')
        
        items_df = pd.read_excel(item_detail_file_path, index_col=False)

        self.load_find_configs()
        
        for config in self.config_list:
            # try:
            if config['enable'] == False:
                continue
            export_name = config['export_name']
            check_conditions = config['check_conditions']
            self.logger.info('Checking conditions: %s'%export_name)

            items_found = self.find_item_matches(items_df, check_conditions)
            items_found.to_excel(os.path.join(self.debug_path, f'items_found.xlsx'), index=False)
            self.export_to_excel(items_found, config['export_mapping'], output_file=os.path.join(self.work_path, f'{export_name}.xlsx'))
            # except Exception as e:
            #     # 处理主逻辑异常
            #     self.logger.error(f"错误: 应用条件时发生异常 - {str(e)}")

    def load_find_configs(self):
        config_path = os.path.join(os.path.dirname(__file__), 'filter_on')
        # 检查目录是否存在
        if os.path.exists(config_path) and os.path.isdir(config_path):
            # 遍历目录中的所有项
            for filename in os.listdir(config_path):
                # 检查文件是否为JSON文件
                if filename.endswith('.json'):
                    file_path = os.path.join(config_path, filename)
                    # 检查是否为文件（不是目录）
                    if os.path.isfile(file_path):
                        self.logger.debug(f"读取配置: {file_path}")
                        with open(file_path, 'r') as f:
                            data = json.load(f)
                            data['export_name'] = filename[:-5]
                            data['enable'] = True
                            self.config_list.append(data)
        else:
            self.logger.error(f"目录不存在: {config_path}")


    def export_to_excel(self, dataframe, export_mapping, output_file='output.xlsx'):
        """
        根据指定的映射关系将DataFrame导出到Excel

        """

        # 创建空的DataFrame用于导出
        export_df = pd.DataFrame()
        
        # 根据映射关系填充导出DataFrame
        for excel_col, df_col in export_mapping.items():
            if isinstance(df_col, list):
                export_df[excel_col] = dataframe[df_col].fillna(0).sum(axis=1)
                continue

            # 直接字段映射
            if df_col in dataframe.columns:
                export_df[excel_col] = dataframe[df_col]
                continue
        
        # 导出到Excel
        # export_df.to_excel(output_file, index=False)

        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            export_df.to_excel(writer, sheet_name='Sheet1', index=False)
            
            # 获取工作簿和工作表对象
            workbook = writer.book
            worksheet = writer.sheets['Sheet1']
            
            # 获取数据的最大列索引和最大行索引
            max_col = export_df.shape[1]
            max_row = export_df.shape[0]

            # 添加筛选器，范围从 A1 到最后一列最后一行
            col_letter = get_excel_column_letter(max_col)
            worksheet.auto_filter.ref = f'A1:{col_letter}{max_row + 1}'

            # 自动调整列宽
            for column_cells in worksheet.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                adjusted_width = (length + 2) * 1.2  # 适当增加宽度
                column_letter = get_excel_column_letter(column_cells[0].column)  # 列索引转字母
                worksheet.column_dimensions[column_letter].width = min(adjusted_width, 30)  # 限制最大宽度

            # 设置自动换行和垂直对齐为居中
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, max_col=worksheet.max_column):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='center')

            # 居中对齐所有单元格（包括表头）
            for row in worksheet.iter_rows(min_row=1, max_row=max_row + 1, min_col=1, max_col=max_col):
                for cell in row:
                    cell.font = Font(name='InconsolataLGC NF', size=12)  # 设置字体
                    cell.alignment = Alignment(horizontal='center', vertical='center')

        self.logger.info(f"数据已成功导出到 {output_file}")
        return export_df
    


    def find_item_matches(self, item_list, check_conditions):
        # 动态应用条件
        def apply_condition(df, condition):
            paths = condition["path"]
            op = condition["operator"]
            if "threshold" in condition:
                threshold = condition["threshold"]
            # threshold = condition["threshold"]

            # 初始化结果掩码
            result_mask = pd.Series([False] * len(df))

            if isinstance(paths, str):
                paths = [paths]

            for path in paths:
                if path not in df.columns:
                    raise ValueError(f"列不存在: {path}")
                
                # 特殊处理：判断空值（NaN）
                if op == "isnull":
                    result_mask |= df[path].isna()
                    continue  # 跳过后续逻辑
                elif op == "notnull":
                    result_mask |= df[path].notna()
                    continue  # 跳过后续逻辑
                elif op == "in":
                    result_mask |= df[path].isin(threshold)
                    continue
                
                series = df[path]
            
                if op == ">":
                    current_mask = series > threshold
                elif op == "<":
                    current_mask = series < threshold
                elif op == ">=":
                    current_mask = series >= threshold
                elif op == "<=":
                    current_mask = series <= threshold
                elif op == "==":
                    current_mask = (series == threshold) | (series == threshold.lower())
                elif op == "!=":
                    current_mask = series != threshold
                else:
                    raise ValueError(f"Unsupported operator: {op}")
                
                # 更新结果掩码（任意列满足条件即可）
                result_mask |= current_mask

            return result_mask

        def apply_conditions(df, config):
            if "logic" in config:
                logic = config["logic"].upper()
                sub_masks = [apply_conditions(df, cond) for cond in config["conditions"]]
                
                if logic == "AND":
                    return pd.Series([all(m) for m in zip(*sub_masks)], index=df.index)
                elif logic == "OR":
                    return pd.Series([any(m) for m in zip(*sub_masks)], index=df.index)
                else:
                    raise ValueError(f"Unsupported logic: {logic}")
            else:
                return apply_condition(df, config)

        mask = apply_conditions(item_list, check_conditions)
        result = item_list[mask]
        return result

