'''
@Description: In User Settings Edit
@Author: your name
@Date: 2019-05-28 19:50:48
@LastEditTime: 2019-09-27 14:38:33
@LastEditors: Please set LastEditors
'''
#!/usr/bin/env python3
#coding=utf-8

from openpyxl import Workbook
import glob
import csv

caredFiledsList = ('mainLoop', 'frameIndex', 'viewName', 'textureMem', 'luaMem')
caredFieldsSet = set(caredFiledsList)
caredFieldsIndexDict = {}
itemPrefix = "[Performance]"

frameInfosDict = {}
frameInfos = []
frozenFrameIndexSet = set()


with open('vega.FrameMonitor.log', newline='') as csvfile:
    rawFrameInfos = csv.reader(csvfile, delimiter=',', quotechar='|')
    firstCSVRow = True
    for row in rawFrameInfos:
        if firstCSVRow:
            firstCSVRow = False
            for i in range(len(row)):
                if row[i] in caredFieldsSet:
                    caredFieldsIndexDict[row[i]] = i
        else:
            frameInfo = {}
            for (k,v) in caredFieldsIndexDict.items():
                frameInfo[k] = row[v]
            
            try:
                if float(frameInfo['mainLoop']) > 0:
                    frozenFrameIndexSet.add(float(frameInfo['frameIndex']))
                    frameInfosDict[frameInfo['frameIndex']] = frameInfo
                    frameInfos.append(frameInfo)
            except:
                pass

# print(frameInfos)
wb = Workbook()

wb.active.title = 'FrozenFrames'
for i in range(len(caredFiledsList)):
    caredField = caredFiledsList[i]
    wb.active.cell(row=1, column=i+1, value=caredField)

for i in range(len(frameInfos)):
    frameInfo = frameInfos[i]
    for j in range(len(caredFiledsList)):
        wb.active.cell(row=i+2, column=j+1, value=frameInfo[caredFiledsList[j]])

frozenFrameInfoDict = {}

listglob = glob.glob("./vega.profile*.log")

profilerItemNameList = ('time', 'stamp', 'frame', 'detail')

for file in listglob:
    f = open(file,"r")  
    lines = f.readlines()
    profilerName = None
    flag = False

    for line in lines:
        # print(line)
        if line.startswith(itemPrefix):
            profilerName = line[len(itemPrefix):].strip()
            flag = False
        elif line.startswith('---------------------'):
            profilerName = None
        elif profilerName != None:

            if not flag:
                flag = True
                continue

            cols = line.split('    ')
            # print(cols)
            emptyColNum = 0
            row = {}
            for index in range(len(cols)):
                col = cols[index].strip()
                if col != None and col != "" and col != ' ':
                    realIndex = index - emptyColNum
                    assert realIndex>=0 and realIndex<len(profilerItemNameList), 'realIndex is {0}, emptyColNum is {1}, col is {2}, ascii is {3}'.format(realIndex, emptyColNum, col, ord(col))

                    colValue = None
                    try:
                        colValue = float(col)
                    except:
                        colValue = col
                    row[profilerItemNameList[realIndex]] = colValue
                else:
                    emptyColNum = emptyColNum + 1

            frozenFrameInfo = None
            if row['frame'] in frozenFrameInfoDict:
                frozenFrameInfo = frozenFrameInfoDict[row['frame']]
            elif row['frame'] in frozenFrameIndexSet:
                frameInfo = frameInfosDict[str(int(row['frame']))]
                totalTime = float(frameInfo['mainLoop'])
                viewName = frameInfo['viewName']
                frozenFrameInfo = {'total':totalTime, 'viewName':viewName, 'frame':row['frame'], 'recordsInfoDict':{}}
                frozenFrameInfoDict[row['frame']] = frozenFrameInfo

            if frozenFrameInfo != None:
                insertedFlag = False
                recordsInfo = None
                if profilerName in frozenFrameInfo['recordsInfoDict']:
                    recordsInfo = frozenFrameInfo['recordsInfoDict'][profilerName]
                else:
                    recordsInfo = {'total':0, 'name':profilerName, 'records':[]}
                    frozenFrameInfo['recordsInfoDict'][profilerName] = recordsInfo

                recordsInfo['total'] = recordsInfo['total'] + row['time']
                records = recordsInfo['records']
                for i in range(len(records)):
                    if row['time'] > records[i]['time']:
                        records.insert(i, row)
                        insertedFlag = True
                        break
                
                if not insertedFlag:
                    records.append(row)

# print(frozenFrameInfoDict)
frozenFrameInfoList = list(frozenFrameInfoDict.values())

def frozenFrameInfoSortKey(frozenFrameInfo):
    return frozenFrameInfo['frame']

def recordsInfoSortKey(recordsInfo):
    return recordsInfo['total']

frozenFrameInfoList.sort(key=frozenFrameInfoSortKey)

for frozenFrameInfo in frozenFrameInfoList:
    
    ws = wb.create_sheet('{0}-{1}'.format(frozenFrameInfo['viewName'], int(frozenFrameInfo['frame'])))
    recordsInfoList = list(frozenFrameInfo['recordsInfoDict'].values())
    recordsInfoList.sort(key=recordsInfoSortKey, reverse=True)
    for i in range(len(recordsInfoList)):
        recordsInfo = recordsInfoList[i]
        ws.cell(row=1, column=i*2+1, value='{0}'.format(recordsInfo['name']))
        ws.cell(row=1, column=i*2+2, value=round(recordsInfo['total'], 3))
        for j in range(len(recordsInfo['records'])):
            record = recordsInfo['records'][j]
            if 'detail' in record:
                ws.cell(row=j+2, column=i*2+1, value='{0}'.format(record['detail']))
                ws.cell(row=j+2, column=i*2+2, value=round(record['time'], 3))

wb.save('vega.statistics.xlsx')