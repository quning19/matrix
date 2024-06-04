#coding=utf8

import os
if os.name != 'nt':
    import sh
import sys
import click
import json
import logging
import openpyxl

delay_info = None
confirm_info = None
excel_file = None

@click.command()
@click.option('-i', '--input', required = True, help=u'log文件路径')
@click.option('-o', '--output', required = True, help=u'excel文件路径')
def run(**options):
    '''导出配置资源, 需要配置路径'''
    print(options['input'])
    print(options['output'])
    global delay_info, confirm_info, excel_file
    excel_file = openpyxl.load_workbook(filename = options['output'], read_only=False, data_only=False)
    read_all_confirm_info(options['input'])
    check_delay_confirmed()
    count_confirms()
    count_frames()
    excel_file.save(options['output'])


def count_confirms():
    global delay_info, confirm_info
    count_label= []
    tcp_counts = []
    udp_counts = []
    tcp_total_delay = 1
    tcp_total_delay_count = 1
    udp_total_delay = 1
    udp_total_delay_count = 1

    def add_to_labels(count_label, delta, count):
        last_one = 20
        if len(count_label) > 0:
            last_one = count_label[len(count_label) - 1]
        for i in range(count):
            count_label.append((i + 1) * delta + last_one)

    add_to_labels(count_label, 10, 8)
    add_to_labels(count_label, 20, 5)
    add_to_labels(count_label, 30, 5)
    add_to_labels(count_label, 50, 5)
    add_to_labels(count_label, 200, 8)
    add_to_labels(count_label, 100000000, 1)

    for i in range(len(count_label)):
        tcp_counts.append(0)
        udp_counts.append(0)

    for key in confirm_info:
        delay = confirm_info[key]['delay']
        if delay < 5000:
            if confirm_info[key]['type'] == 'TCP':
                tcp_total_delay_count += 1
                tcp_total_delay += delay
            if confirm_info[key]['type'] == 'UDP':
                udp_total_delay_count += 1
                udp_total_delay += delay
        for i in range(len(count_label)):
            if delay <= count_label[i]:
                if confirm_info[key]['type'] == 'TCP':
                    tcp_counts[i] += 1
                if confirm_info[key]['type'] == 'UDP':
                    udp_counts[i] += 1
                break

    worksheet = excel_file.get_sheet_by_name(excel_file.sheetnames[0])
    worksheet.cell(row = 1, column = 1).value = "count"
    worksheet.cell(row = 1, column = 2).value = "tcp"
    worksheet.cell(row = 1, column = 3).value = "udp"
    for i in range(len(count_label)):
        worksheet.cell(row = i + 2, column = 1).value = count_label[i]
        worksheet.cell(row = i + 2, column = 2).value = tcp_counts[i]
        worksheet.cell(row = i + 2, column = 3).value = udp_counts[i]

    worksheet.cell(row = 1, column = 4).value = "tcp avarge"
    worksheet.cell(row = 2, column = 4).value = tcp_total_delay/tcp_total_delay_count
    worksheet.cell(row = 1, column = 5).value = "udp avarge"
    worksheet.cell(row = 2, column = 5).value = udp_total_delay/udp_total_delay_count

def count_frames():
    global delay_info, confirm_info
    count_label= []
    tcp_counts = []
    udp_counts = []
    tcp_total_delay = 1
    tcp_total_delay_count = 1
    udp_total_delay = 1
    udp_total_delay_count = 1

    def add_to_labels(count_label, delta, count):
        last_one = 0
        if len(count_label) > 0:
            last_one = count_label[len(count_label) - 1]
        for i in range(count):
            count_label.append((i + 1) * delta + last_one)

    add_to_labels(count_label, 1, 6)
    add_to_labels(count_label, 2, 2)
    add_to_labels(count_label, 5, 4)
    add_to_labels(count_label, 10, 2)
    # add_to_labels(count_label, 200, 1)
    add_to_labels(count_label, 10000, 1)

    for i in range(len(count_label)):
        tcp_counts.append(0)
        udp_counts.append(0)

    for key in delay_info:
        delay = delay_info[key]['delay']
        if delay_info[key]["confirmed"]:
            if delay_info[key]['type'] == 'TCP':
                tcp_total_delay_count += 1
                tcp_total_delay += delay
            if delay_info[key]['type'] == 'UDP':
                udp_total_delay_count += 1
                udp_total_delay += delay
            for i in range(len(count_label)):
                if delay <= count_label[i]:
                    if delay_info[key]['type'] == 'TCP':
                        tcp_counts[i] += 1
                    if delay_info[key]['type'] == 'UDP':
                        udp_counts[i] += 1
                    break

    worksheet = excel_file.get_sheet_by_name(excel_file.sheetnames[0])
    worksheet.cell(row = 1, column = 7).value = "count frames"
    worksheet.cell(row = 1, column = 8).value = "tcp frames"
    worksheet.cell(row = 1, column = 9).value = "udp frames"
    for i in range(len(count_label)):
        worksheet.cell(row = i + 2, column = 7).value = count_label[i]
        worksheet.cell(row = i + 2, column = 8).value = tcp_counts[i]
        worksheet.cell(row = i + 2, column = 9).value = udp_counts[i]

    worksheet.cell(row = 1, column = 10).value = "tcp frame avarge"
    worksheet.cell(row = 2, column = 10).value = tcp_total_delay * 1.0 / tcp_total_delay_count
    worksheet.cell(row = 1, column = 11).value = "udp frame avarge"
    worksheet.cell(row = 2, column = 11).value = udp_total_delay * 1.0 / udp_total_delay_count

def check_delay_confirmed():
    global delay_info, confirm_info
    for key in delay_info:
        delay_info[key]['confirmed'] = key in confirm_info
        # if not delay_info[key]['confirmed']:
        #     print delay_info[key]



def read_all_confirm_info(fileName):
    global delay_info, confirm_info
    delay_info={}
    confirm_info = {}

    #print fileName
    if os.path.exists(fileName):
        logFile = open(fileName)
        while True:
            line = logFile.readline()
            if len(line) == 0:
                break
            line = line[:-1].split(" : ")[1].split(',')
            if line[0][-10:] == 'PVPCONFIRM':
                confirm_info[line[1]+'|'+line[2]+'|'+line[3]] = {"delay" : int(line[4]), 'type':line[0][:3]}
            if line[0][-8:] == "PVPDELAY":
                delay_info[line[1]+'|'+line[2]+'|'+line[3]] = {"delay" : int(line[4]), 'type':line[0][:3]}











 #
